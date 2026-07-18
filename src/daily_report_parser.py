"""
daily_report_parser.py — parse a PGCB/NLDC daily report workbook (QF-LDC forms)
into canonical CSVs under data/pipeline/canonical/daily-reports/<data-date>/.

The workbook is the operational daily report published by PGCB's National Load
Dispatch Centre (sheets: Forecast, GenLog, P1, P2, P3, P4, L-Curve, En-Curve,
EWIC, Voltage). The report is dated one day after the actuals it contains; the
output folder is named after the ACTUALS date (ISO format).

Usage:
    python src/daily_report_parser.py <path-to-Daily_Report_xxx.xlsx>

Outputs (one folder per data date):
    plant_generation.csv   per-plant hourly MW + daily total kWh (GenLog)
    plant_attributes.csv   fuel/producer/capacity/remarks (Forecast + P2)
    hourly_system.csv      hourly generation / load shed / demand (P4)
    fuel_curve.csv         half-hourly generation by fuel bucket (En-Curve)
    substation_peaks.csv   per-substation daily peak load + time (P3)
    area_peaks.csv         zone/area peak loads (P3 header block)
    ewic_flows.csv         hourly East-West interconnector flows (EWIC)
    summary.csv            key system metrics incl. production cost (P1)
"""

from __future__ import annotations

import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).parent.parent
OUT_BASE = ROOT / "data" / "pipeline" / "canonical" / "daily-reports"

# GenLog columns that are aggregates, not plants
GENLOG_AGGREGATES = {"Eastern Total", "Western Total", "National Grid Total", "Water Level"}

# GenLog columns that are cross-border energy purchases, not fuel-burning plants
IMPORT_NAME_PATTERNS = [
    re.compile(r"import", re.I),          # "Import (Tripura)"
    re.compile(r"^bheramara\s*\(\s*hvdc", re.I),  # "Bheramara (HVDC )"
    re.compile(r"^hvdc", re.I),           # "HVDC(Nepal)"
    re.compile(r"adani", re.I),           # "Adani Power Jharkhanda Ltd."
]


def _clean(v):
    """Collapse internal whitespace/newlines of a header cell."""
    if v is None:
        return None
    return re.sub(r"\s+", " ", str(v)).strip()


def _num(v):
    """Best-effort numeric conversion; N/A and text become None."""
    if v is None or isinstance(v, str) and not re.search(r"\d", v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


def _time_str(v):
    """Normalize a time cell (datetime.time, '19:30', '19:30:00') to HH:MM."""
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%H:%M")
    s = str(v).strip()
    m = re.match(r"^(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return s if s and s.upper() != "N/A" else None


def _parse_date(v) -> date | None:
    """Parse a report date cell: datetime, 'DD-MM-YYYY', or 'DD/MM/YYYY'."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", str(v))
    if m:
        d, mth, y = (int(g) for g in m.groups())
        return date(y, mth, d)
    return None


def find_labeled_date(ws, max_row=10, max_col=30) -> date | None:
    """Scan the top of a sheet for a 'Date' label and parse the cell right of it."""
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None and re.match(r"^\s*date\b", str(cell.value), re.I):
                for offset in (1, 2):
                    d = _parse_date(ws.cell(row=cell.row, column=cell.column + offset).value)
                    if d:
                        return d
    return None


# ── GenLog ─────────────────────────────────────────────────────────────────

def parse_genlog(ws) -> pd.DataFrame:
    """Per-plant hourly MW and daily total kWh.

    Layout: plant names row 11, rated capacity ('102.000 MW') row 12, hour rows
    13..37 (00:00..23:00 with an extra 19:30), 'Total KWH' row 38. Aggregate
    columns (Eastern/Western/National totals, Water Level) are excluded; area is
    inferred from position relative to the Eastern Total column.
    """
    NAME_ROW, CAP_ROW = 11, 12

    # locate the hour rows and the Total KWH row by their col-A labels
    hour_rows: list[tuple[int, str]] = []
    total_row = None
    for r in range(CAP_ROW + 1, ws.max_row + 1):
        label = _clean(ws.cell(row=r, column=1).value)
        if label is None:
            continue
        if re.match(r"^total\s*kwh", label, re.I):
            total_row = r
            break
        t = _time_str(label)
        if t and re.match(r"^\d{2}:\d{2}$", t):
            hour_rows.append((r, t))
    if total_row is None:
        raise ValueError("GenLog: 'Total KWH' row not found")

    # locate plant columns and the Eastern Total boundary
    eastern_total_col = None
    cols: list[tuple[int, str]] = []
    for c in range(2, ws.max_column + 1):
        name = _clean(ws.cell(row=NAME_ROW, column=c).value)
        if not name:
            continue
        if name in GENLOG_AGGREGATES:
            if name == "Eastern Total":
                eastern_total_col = c
            continue
        cols.append((c, name))

    records = []
    for c, name in cols:
        cap = _num(str(ws.cell(row=CAP_ROW, column=c).value or "").replace("MW", ""))
        rec = {
            "name": name,
            "capacity_mw": cap,
            "area": "Eastern" if (eastern_total_col and c < eastern_total_col) else "Western",
            "is_import": any(p.search(name) for p in IMPORT_NAME_PATTERNS),
        }
        for r, t in hour_rows:
            rec[f"mw_{t.replace(':', '')}"] = _num(ws.cell(row=r, column=c).value)
        rec["total_kwh"] = _num(ws.cell(row=total_row, column=c).value)
        records.append(rec)

    df = pd.DataFrame(records)
    # duplicate names (e.g. two 583 MW Meghnaghat Summit columns) get a suffix
    dup = df["name"].duplicated(keep=False)
    df.loc[dup, "name"] = (
        df.loc[dup, "name"] + " #" + (df.loc[dup].groupby("name").cumcount() + 1).astype(str)
    )
    return df


# ── Forecast + P2 → plant attributes ───────────────────────────────────────

def parse_forecast(ws) -> pd.DataFrame:
    """Fuel/producer/capacity/remarks per plant from the Forecast sheet.

    Data rows start at 11; a row is a plant row when the name (col 2) and fuel
    (col 3) are both present. The half-hourly forecast block at the sheet tail
    has neither, so it is skipped naturally.
    """
    records = []
    for r in range(11, ws.max_row + 1):
        name = _clean(ws.cell(row=r, column=2).value)
        fuel = _clean(ws.cell(row=r, column=3).value)
        if not name or not fuel or "total" in name.lower():
            continue
        records.append({
            "name": name,
            "fuel": fuel,
            "producer": _clean(ws.cell(row=r, column=4).value),
            "unit_config": _clean(ws.cell(row=r, column=5).value),
            "installed_mw": _num(ws.cell(row=r, column=6).value),
            "present_mw": _num(ws.cell(row=r, column=7).value),
            "actual_day_peak_mw": _num(ws.cell(row=r, column=8).value),
            "actual_evening_peak_mw": _num(ws.cell(row=r, column=9).value),
            "forecast_day_peak_mw": _num(ws.cell(row=r, column=10).value),
            "forecast_evening_peak_mw": _num(ws.cell(row=r, column=11).value),
            "remarks": _clean(ws.cell(row=r, column=14).value),
        })
    return pd.DataFrame(records)


def parse_p2(ws) -> pd.DataFrame:
    """Per-plant evening-peak MW and daily energy kWh from P2 (cross-check)."""
    records = []
    for r in range(11, ws.max_row + 1):
        name = _clean(ws.cell(row=r, column=3).value)
        if not name or "total" in name.lower():
            continue
        energy = _num(ws.cell(row=r, column=9).value)
        if energy is None and _num(ws.cell(row=r, column=8).value) is None:
            continue  # header/signature rows
        records.append({
            "name": name,
            "p2_evening_peak_mw": _num(ws.cell(row=r, column=8).value),
            "p2_energy_kwh": energy,
            "p2_remarks": _clean(ws.cell(row=r, column=10).value),
        })
    return pd.DataFrame(records)


# ── P1 summary ─────────────────────────────────────────────────────────────

P1_METRICS = {
    "day_peak_generation_mw": r"day peak generation",
    "day_peak_demand_mw": r"day peak demand",
    "evening_peak_generation_mw": r"evening peak gener",
    "evening_peak_demand_mw": r"evening peak deman",
    "min_generation_mw": r"minimum generation",
    "max_generation_mw": r"maximum generation",
    "energy_generated_mkwh": r"energy generated",
    "energy_unserved_mkwh": r"energy unserve",
    "energy_demand_mkwh": r"energy demand",
    "max_temperature_c": r"maximum temperatur",
    "gas_supplied_mmcfd": r"total gas supplied",
    "production_cost_tk_per_kwh": r"production cost",
}


def parse_p1(ws) -> pd.DataFrame:
    """Key system metrics from the P1 summary block (label → first number right)."""
    found = {}
    for row in ws.iter_rows(min_row=1, max_row=30, max_col=15):
        for cell in row:
            if cell.value is None or not isinstance(cell.value, str):
                continue
            label = _clean(cell.value).lower()
            for key, pat in P1_METRICS.items():
                if key not in found and re.search(pat, label):
                    for c in range(cell.column + 1, cell.column + 10):
                        v = _num(ws.cell(row=cell.row, column=c).value)
                        if v is not None:
                            found[key] = v
                            break
    return pd.DataFrame(sorted(found.items()), columns=["metric", "value"])


# ── P3 substation & area peaks ─────────────────────────────────────────────

def parse_p3(ws) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Substation daily peak loads (4 side-by-side column groups) + zone totals."""
    # zone/area block, rows 8-14: labels in col 2 (east) and col 8 (west)
    areas = []
    for r in range(8, 15):
        for label_col, val_cols in ((2, range(3, 8)), (8, range(9, 14))):
            label = _clean(ws.cell(row=r, column=label_col).value)
            if not label:
                continue
            val = next((v for c in val_cols
                        if (v := _num(ws.cell(row=r, column=c).value)) is not None), None)
            if val is not None:
                areas.append({"area": re.sub(r"^[ivx()\s.]+", "", label), "peak_mw": val})
    area_df = pd.DataFrame(areas)

    subs = []
    for r in range(17, ws.max_row + 1):
        for base in (2, 6, 10, 14):
            sl = _num(ws.cell(row=r, column=base).value)
            name = _clean(ws.cell(row=r, column=base + 1).value)
            if sl is None or not name:
                continue
            subs.append({
                "sl": int(sl),
                "substation": name,
                "peak_mw": _num(ws.cell(row=r, column=base + 2).value),
                "peak_time": _time_str(ws.cell(row=r, column=base + 3).value),
            })
    sub_df = pd.DataFrame(subs).sort_values("sl").reset_index(drop=True)
    return sub_df, area_df


# ── P4 hourly system ───────────────────────────────────────────────────────

def parse_p4(ws) -> pd.DataFrame:
    records = []
    for r in range(9, ws.max_row + 1):
        t = _time_str(ws.cell(row=r, column=2).value)
        gen = _num(ws.cell(row=r, column=3).value)
        if not t or gen is None:
            continue
        records.append({
            "time": t,
            "generation_mw": gen,
            "load_shed_mw": _num(ws.cell(row=r, column=5).value),
            "demand_mw": _num(ws.cell(row=r, column=7).value),
        })
    return pd.DataFrame(records)


# ── En-Curve fuel mix ──────────────────────────────────────────────────────

def parse_en_curve(ws) -> pd.DataFrame:
    """Half-hourly generation by fuel bucket. Header row 3, data rows 4+."""
    HEADER_ROW = 3
    cols = []
    for c in range(2, 20):
        name = _clean(ws.cell(row=HEADER_ROW, column=c).value)
        if name:
            cols.append((c, re.sub(r"\W+", "_", name.lower()).strip("_")))
    records = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        t = _time_str(ws.cell(row=r, column=1).value)
        if not t or not re.match(r"^\d{2}:\d{2}$", t):
            continue
        rec = {"time": t}
        for c, name in cols:
            rec[name] = _num(ws.cell(row=r, column=c).value)
        records.append(rec)
    return pd.DataFrame(records)


# ── EWIC corridor flows ────────────────────────────────────────────────────

def parse_ewic(ws) -> pd.DataFrame:
    corridors = ["ishurdi_ghorashal", "sirajganj_sreepur", "gopalganj_aminbazar"]
    records = []
    for r in range(9, ws.max_row + 1):
        t = _time_str(ws.cell(row=r, column=1).value)
        if not t or not re.match(r"^\d{2}:\d{2}$", t):
            continue
        rec = {"time": t}
        for i, corridor in enumerate(corridors):
            rec[f"{corridor}_mw"] = _num(ws.cell(row=r, column=2 + 2 * i).value)
            rec[f"{corridor}_mvar"] = _num(ws.cell(row=r, column=3 + 2 * i).value)
        records.append(rec)
    return pd.DataFrame(records)


# ── main ───────────────────────────────────────────────────────────────────

def main() -> None:
    if len(sys.argv) != 2:
        sys.exit("Usage: python src/daily_report_parser.py <Daily_Report.xlsx>")
    xlsx = Path(sys.argv[1])
    if not xlsx.exists():
        sys.exit(f"Not found: {xlsx}")

    print(f"Reading {xlsx.name} ...")
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # actuals date comes from an actuals sheet (P4/P2/GenLog), not the cover date
    data_date = None
    for sheet in ("P4", "P2", "GenLog"):
        if sheet in wb.sheetnames:
            data_date = find_labeled_date(wb[sheet], max_row=10, max_col=30)
            if data_date:
                break
    if data_date is None:
        sys.exit("Could not find the actuals date on P4/P2/GenLog")

    out_dir = OUT_BASE / data_date.isoformat()
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"Data date: {data_date.isoformat()}  ->  {out_dir.relative_to(ROOT)}/")

    gen = parse_genlog(wb["GenLog"])
    forecast = parse_forecast(wb["Forecast"])
    p2 = parse_p2(wb["P2"])
    attrs = forecast.merge(p2, on="name", how="outer")
    summary = parse_p1(wb["P1"])
    hourly = parse_p4(wb["P4"])
    fuel_curve = parse_en_curve(wb["En-Curve"])
    subs, areas = parse_p3(wb["P3"])
    ewic = parse_ewic(wb["EWIC"])

    for fname, df in {
        "plant_generation.csv": gen,
        "plant_attributes.csv": attrs,
        "summary.csv": summary,
        "hourly_system.csv": hourly,
        "fuel_curve.csv": fuel_curve,
        "substation_peaks.csv": subs,
        "area_peaks.csv": areas,
        "ewic_flows.csv": ewic,
    }.items():
        df.to_csv(out_dir / fname, index=False)
        print(f"  {fname:<24} {len(df):>4} rows")

    # ── consistency checks ────────────────────────────────────────────────
    print("\nConsistency checks:")
    total_kwh = gen["total_kwh"].sum()
    p1_mkwh = summary.set_index("metric")["value"].get("energy_generated_mkwh")
    print(f"  GenLog plant total : {total_kwh / 1e6:12.3f} MkWh")
    if p1_mkwh is not None:
        dev = 100 * (total_kwh / 1e6 - p1_mkwh) / p1_mkwh
        print(f"  P1 energy generated: {p1_mkwh:12.3f} MkWh   (deviation {dev:+.3f} %)")

    both = gen.merge(p2, on="name", how="inner").dropna(subset=["total_kwh", "p2_energy_kwh"])
    if len(both):
        diff = (both["total_kwh"] - both["p2_energy_kwh"]).abs()
        agree = (diff < 1e3) | (diff / both["p2_energy_kwh"].clip(lower=1) < 0.01)
        print(f"  GenLog vs P2 energy: {agree.sum()}/{len(both)} matched plants agree within 1%")

    unmatched = gen.loc[~gen["is_import"] & ~gen["name"].isin(attrs["name"]), "name"]
    print(f"  GenLog plants without a Forecast/P2 attribute row: {len(unmatched)}")
    if len(unmatched):
        print("    (fuel for these must come from name suffix or the estimator's lookups)")

    print("\nDone.")


if __name__ == "__main__":
    main()
