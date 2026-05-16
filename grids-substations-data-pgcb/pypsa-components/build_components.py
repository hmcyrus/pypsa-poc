#!/usr/bin/env python3
"""
Build PyPSA Bus and Line components from PGCB grid data.

Sources:
  - Grid.xlsx  : substation list (colors encode voltage level)
  - grids-formatted.csv : transmission line list (has explicit Voltage Level column)

Output:
  - pypsa_buses.csv
  - pypsa_lines.csv
"""

import re
import csv
import sys
from pathlib import Path

import openpyxl
import pandas as pd

DATA_DIR = Path(__file__).parent.parent
OUT_DIR = Path(__file__).parent


# ---------------------------------------------------------------------------
# 1.  SUBSTATION → BUS parsing (Grid.xlsx)
# ---------------------------------------------------------------------------

# Row ranges (1-indexed) and their inferred voltage level.
# Determined by openpyxl color scan of Grid.xlsx.
SECTION_RULES = [
    # (first_row, last_row, color_hint, voltage_kV)
    (5,   6,   "FFFFFFCC", 400),   # BIPTC 400 kV power plant substations
    (7,  19,   "FFCCCCFF", 400),   # Main 400 kV substations
    (20, 23,   "FFFFFF99", None),  # Mixed – handled per-name below
    (24, 64,   "FFC0C0C0", 230),   # 230 kV substations
    (65, 70,   "FFCCCCFF", 230),   # 230 kV bulk consumers
    (71, 9999, "FFCCCCFF", 132),   # 132 kV distribution substations
]

# In the FFFFFF99 section, these names are SECONDARY (230 kV) buses because
# the primary (400 kV) bus already exists in the FFCCCCFF-400 section.
YELLOW_SECONDARY_230 = {"Gopalganj (N)", "Kaliakoir"}


def _voltage_for_row(row_idx: int, name: str) -> int:
    for first, last, _color, voltage in SECTION_RULES:
        if first <= row_idx <= last:
            if voltage is not None:
                return voltage
            # FFFFFF99 mixed section
            return 230 if name in YELLOW_SECONDARY_230 else 400
    return 132  # fallback


def _clean(val) -> str:
    if val is None:
        return ""
    return str(val).replace("\n", " ").strip()


def parse_substations() -> pd.DataFrame:
    wb = openpyxl.load_workbook(DATA_DIR / "Grid.xlsx")
    ws = wb.active

    rows = []
    seen_names: set[str] = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), 1):
        cells = list(row)
        if len(cells) < 2:
            continue

        sn_val = cells[0].value
        name_val = cells[1].value

        # Skip header / empty rows
        if not isinstance(sn_val, (int, float)):
            continue
        if name_val is None or "Name of Grid" in str(name_val):
            continue

        name = _clean(name_val)
        if not name:
            continue

        transformer = _clean(cells[3].value) if len(cells) > 3 else ""
        capacity    = _clean(cells[5].value) if len(cells) > 5 else ""
        ownership   = _clean(cells[7].value) if len(cells) > 7 else ""
        grid_circle = _clean(cells[8].value) if len(cells) > 8 else ""

        v_nom = _voltage_for_row(row_idx, name)
        bus_name = f"{name}_{v_nom}kV"

        if bus_name not in seen_names:
            seen_names.add(bus_name)
            rows.append({
                "name":              bus_name,
                "substation":        name,
                "v_nom":             v_nom,
                "carrier":           "AC",
                "transformer_detail": transformer,
                "total_capacity_mva": capacity,
                "ownership":         ownership,
                "grid_circle":       grid_circle,
                "implicit":          False,
            })

    # Add implicit secondary buses for substations that appear on lines at a
    # voltage level not documented in the xlsx.
    existing = {(r["substation"], r["v_nom"]) for r in rows}
    name_to_row = {r["substation"]: r for r in rows}

    for sub_name, extra_voltages in IMPLICIT_SECONDARY_BUSES.items():
        base = name_to_row.get(sub_name, {})
        for v in extra_voltages:
            if (sub_name, v) not in existing:
                bus_name = f"{sub_name}_{v}kV"
                if bus_name not in seen_names:
                    seen_names.add(bus_name)
                    rows.append({
                        "name":              bus_name,
                        "substation":        sub_name,
                        "v_nom":             v,
                        "carrier":           "AC",
                        "transformer_detail": base.get("transformer_detail", ""),
                        "total_capacity_mva": "",
                        "ownership":         base.get("ownership", ""),
                        "grid_circle":       base.get("grid_circle", ""),
                        "implicit":          True,
                    })

    # Add completely new buses (appear in lines but not in xlsx)
    for bus_name, v in EXTRA_BUSES.items():
        if bus_name not in seen_names:
            seen_names.add(bus_name)
            # substation name = everything before the last "_<v>kV"
            sub = bus_name.rsplit("_", 1)[0]
            rows.append({
                "name":              bus_name,
                "substation":        sub,
                "v_nom":             v,
                "carrier":           "AC",
                "transformer_detail": "",
                "total_capacity_mva": "",
                "ownership":         "",
                "grid_circle":       "",
                "implicit":          True,
            })

    df = pd.DataFrame(rows)
    return df


# ---------------------------------------------------------------------------
# 2.  NAME NORMALISATION for line-endpoint matching
# ---------------------------------------------------------------------------

# Common spelling/abbreviation mismatches between line names and substation names.
# Keys are what appears in grids-formatted.csv; values are canonical substation names.
NAME_ALIASES: dict[str, str] = {
    # 400 kV
    "Comilla(N)":           "Cumilla (N)",
    "Comilla North":        "Cumilla (N)",
    "Comilla (N)":          "Cumilla (N)",
    "Comilla (S)":          "Cumilla (S)",
    "Cumilla(N)":           "Cumilla (N)",
    "Cumilla(S)":           "Cumilla (S)",
    "Rooppur":              "RNPP",
    "Roopur":               "RNPP",
    "RNPL":                 "RNPP",
    "Amtali":               "Amtali Switching",
    "Amtoli":               "Amtali Switching",
    "Ashuganj(N)":          "Ashuganj (N)",
    "Ashuganj (N)":         "Ashuganj (N)",
    "Bogura(W)":            "Bogura (W)",
    "Bogura(West)":         "Bogura (W)",
    "Gopalganj(N)":         "Gopalganj (N)",
    "Madunaghat(O)":        "Madunaghat",
    "Madunaght(O)":         "Madunaghat",
    # 230 kV
    "Ghorasal":             "Ghorashal",
    "Ghorashal":            "Ghorashal",
    "Ishurdi":              "Ishwardi",
    "Bogra":                "Bogura (S)",
    "Bogura(S)":            "Bogura (S)",
    "Bogura":               "Bogura (S)",
    "Barisal":              "Barishal (N)",
    "Barishal":             "Barishal (N)",
    "Barishal(N)":          "Barishal (N)",
    "Khulna HVDC":          "Bheramara",
    "Khulna-Bheramara HVDC":"Bheramara",
    "Bheramara HVDC":       "Bheramara",
    "Bheramana":            "Bheramara",
    "Bhermara":             "Bheramara",
    "Old Airport":          "Agargaon",
    "Patuakhali":           "Patuakhali",
    "Payra PP":             "Payra",
    "Payra SS":             "Payra",
    "AES, Haripur":         "Haripur",
    "Siddhirganj 210 MW P/S": "Siddhirganj",
    "Sonagazi (EGCB)":      "Mirsarai",
    "Khulna(S)":            "Khulna (S)",
    "Khulna ©":             "Khulna (C)",
    "Khulna(S)":            "Khulna (S)",
    "Mongla- Khulna(S)":    "Mongla",
    # 132 kV
    "Jessore":              "Jashore",
    "Kustia":               "Kushtia",
    "Bheramara HVDC-Bheramara 230": "Bheramara",
    "Joydevpur":            "Joydebpur",
    "Julda":                "Juldah",
    "Raojan":               "Raozan",
    "Niyamatpur":           "Niamatpur",
    "Chowdala":             "Chowdala",
    "TKC":                  "TKCCL",
    "Fenchuganj PS":        "Fenchuganj",
    "Fenchuganj P/S":       "Fenchuganj",
    "B.Baria":              "Brahmanbaria",
    "BograOld":             "Bogura",
    "BograNew":             "Bogura",
    "Shiddhirganj":         "Siddhirganj",
    "Confidence PP":        "Bogura (S)",
    "Confedence PP":        "Bogura (S)",
    "United PP":            "Sherpur",
    "Faridpur PP":          "Faridpur",
    "Amnura PP":            "Amnura",
    "Noapara PP":           "Noapara",
    "Daudkandi PP":         "Daudkandi",
    "Gopalganj PP":         "Gopalganj",
    "Modhumati PP":         "Gopalganj",
    "Sirangj 68MW Solar":   "Sirajganj",
    "Sirajganj Solar":      "Sirajganj",
    "Sirajganj Solar-Sirajganj Grid": "Sirajganj",
    "Sudarganj Solar":      "Rangpur",
    "Sudarganj Solar - Rangpur": "Rangpur",
    "Labanchora PP":        "Goalpara",
    "Dynamics Solar Plant": "Ishwardi",
    "US DK":                "Cox's Bazar",
    "Anandabazar":          "Halishahar",
    "Muktagacha":           "Muktagacha",
    "Mukhtagacha":          "Muktagacha",
    "Ghatail":              "Ghatail",
    "Ghatial":              "Ghatail",
    "Sherpur (Bogura)":     "Sherpur(Bogura)",
    "Chapai-Nawabganj":     "Chapai Nawabganj",
    "Feni 230/132 kV":      "Feni (N)",
    "Feni Chauddagram":     "Chauddagram",
    "Feni Chowmuhani":      "Feni (N)",
    "Dagunbhuyan":          "Daganbhuiyan",
    "Daganbhuiyan":         "Daganbhuiyan",
    "Bhulta (N)":           "Bhulta",
    "Bhulta(N)":            "Bhulta",
    "Pabna 64 MW Solar":    "Pabna",
    "Rajshahi)N)":          "Rajshahi(N)",
    "Sholoshahor":          "Sholoshahar",
    "Chauddagram":          "Chauddagram",
    "Chowddagram":          "Chauddagram",
    "DU":                   "Dhaka University (DU)",
    "Manaknagar":           "Maniknagar",
    "Noagaon":              "Naogaon",
    "Basundhara":           "Bashundhara",
    "Basundhara Cement":    "Bashundhara Cement",
    "Sreepur":              "Sripur",
    "Hairpur":              "Haripur",
    "Postogola":            "Postagola",
    "Sajmasjid":            "Satmasjid",
    "Madertek":             "Madartek",
    "Munsiganj":            "Munshiganj",
    "Bangura":              "Bhangura",
    "Gopalanj":             "Gopalganj",
    "Lalmonirht 2nd circuit": "Lalmonirhat",
    "Rahanpur (LILO point)": "Rahanpur",
    "Barmi (BR Powergen)":  "Barmi",
    "Kodda PP":             "Kodda",
    "Pubail":               "Pubail",
    "Araihazar":            "Araihazar",
    "Sathia":               "Sathia",
    "Sripur":               "Sripur",
    "Bakerganj":            "Bakerganj",
    "Kaliganj":             "Kaliganj",
    "Maheshpur":            "Maheshpur",
    "Barmi":                "Barmi",
    "Gazaria":              "Gazaria",
    "Cantonment":           "Cantonment",
    "Anowara":              "Anowara",
    "Banskhali":            "Banskhali",
    "Monakosa":             "Monakosa",
    "Madunaght":            "Madunaghat",
    "Comilla":              "Cumilla (N)",   # bare "Comilla" → N variant by default
    "Raojan":               "Raozan",
    "Anowara":              "Anowara",
    "Gazaria":              "Gazaria",
    "Sonagazi":             "Sonagazi",
    "Bakerganj":            "Bakerganj",
    "Barmi":                "Barmi",
    "Pubail":               "Pubail",
    "Araihazar":            "Araihazar",
    "Kaliganj":             "Kaliganj",
    "Maheshpur":            "Maheshpur",
    "Cantonment":           "Cantonment",
    "Khula":                "Khulna (S)",    # "Khula 330MW-Khula(S)" → Khulna(S)
    "Khula(S)":             "Khulna (S)",
    "Comilla(S)":           "Cumilla (S)",
    "Barisal(N)":           "Barishal (N)",
    "Lalmonirht":           "Lalmonirhat",
    "Megnaghat Rental PP":  "Meghnaghat",
    "Shiddhirganj desh energy": "Siddhirganj",
    "Bhola":                "Bhola",
    "RPCL":                 "RPCL",
    "Energaon":             "Energaon",
    "Kabir Steel":          "Kabir Steel",
    "RPCL PP":              "Tangail",       # RPCL generator connects via Tangail
    "RPCL":                 "Tangail",       # RPCL power plant at Tangail
    "Megnaghat":            "Meghnaghat",    # spelling variant
    "Megnaghat Rental PP":  "Meghnaghat",
}

# Additional extra buses for remaining unresolved nodes
# (merged into EXTRA_BUSES below)

# Substations that don't exist in Bangladesh's grid model (borders, generators)
# Lines referencing these as endpoints will still be emitted but marked.
OUT_OF_MODEL = {
    "Bangladesh Border(Baharampur)",
    "Bangladesh Border",
    "Bangladesh Border(Baharampur) 2nd",
    "Khula 330MW",          # generator unit, not a substation
    "Barguna PP",
    "Energaon",             # generator near Mongla
    "RPCL",                 # Rooppur Power Company Ltd generator
    "Kabir Steel",          # bulk industrial consumer
}

# Buses that appear in line data but have no entry in the xlsx at all.
# Format: { bus_name : v_nom }
EXTRA_BUSES: dict[str, int] = {
    "Banskhali_400kV":   400,   # 400kV junction on Matarbari–Madunaghat corridor
    "Monakosa_400kV":    400,   # 400kV switching point on Bogura(W)–Rahanpur line
    "Raozan_230kV":      230,   # Raozan power-plant bus (230kV, Chittagong region)
    "Bhola_230kV":       230,   # Bhola island 230kV substation
    "Anowara_230kV":     230,   # Anowara 230kV substation near Chittagong
    "Gazaria_230kV":     230,   # Gazaria 230kV on Kachua–Gazaria line
    "Sonagazi_230kV":    230,   # Sonagazi EGCB 230kV (near Mirsarai)
    "Bakerganj_132kV":   132,   # Bakerganj 132kV, Barisal region
    "Barmi_132kV":       132,   # Barmi 132kV, Sherpur region
    "Pubail_132kV":      132,   # Pubail 132kV, LILO on Ghorashal–Joydevpur
    "Araihazar_132kV":   132,   # Araihazar 132kV
    "Kaliganj_132kV":    132,   # Kaliganj 132kV
    "Maheshpur_132kV":   132,   # Maheshpur 132kV
    "Cantonment_132kV":  132,   # Cantonment (Dhaka) 132kV
    "Madunaght_230kV":   230,   # old Madunaghat spelling variant
    "Kabir Steel_132kV": 132,   # bulk industrial consumer near Baroaulia
    "Bangladesh Border_400kV": 400,  # India interconnection point
    "Khula 330MW_230kV": 230,   # 330 MW generator bus at Khulna
    "Anowara_230kV":     230,   # Anowara 230kV (already listed, kept for dedup safety)
    "Energaon_132kV":    132,   # Energaon generation node near Mongla
}

# Substations that need implicit buses at an additional voltage level because
# the transmission line data references them there but the xlsx only lists them
# at the primary voltage.  Format: { canonical_name : [extra_voltage, ...] }
IMPLICIT_SECONDARY_BUSES: dict[str, list[int]] = {
    "Korerhat":      [230],   # 400kV SS but has 230kV lines (Korerhat-Chowmuhani)
    "Bibiyana":      [230],   # 400kV SS but has 230kV lines (Fenchuganj-Bibiyana)
    "RNPP":          [230],   # 400kV SS but has 230kV line (Rooppur-Baghabari)
    "Bheramara":     [400],   # 230kV SS but HVDC lines are listed at 400kV
    "Baghabari":     [132],   # 230kV SS but 132kV lines reference it
    "Payra":         [230],   # 400kV SS but 230kV line (Patuakhali-Payra)
    "Patuakhali":    [230],   # 132kV SS but 230kV line exists
    "Mongla":        [230],   # 132kV SS but 230kV line (Mongla-Khulna(S))
    "Cumilla (N)":   [400],   # 230kV SS but 400kV line (Comilla(N)-Bangladesh Border)
    "Ashuganj (N)":  [230],   # 400kV SS but 230kV line "Ghorasal-Ashuganj" may reach it
    "Bogura (W)":    [230],   # 400kV SS but LILO line at 230kV references it
}


def normalise_bus_name(raw: str) -> str:
    """Strip extra whitespace/punctuation and apply alias map."""
    raw = raw.strip().rstrip(".")
    return NAME_ALIASES.get(raw, raw)


# ---------------------------------------------------------------------------
# 3.  LINE PARSING (grids-formatted.csv)
# ---------------------------------------------------------------------------

VOLTAGE_MAP = {
    "400 kV": 400,
    "400kV":  400,
    "230 kV": 230,
    "230kV":  230,
    "132 kV": 132,
    "132kV":  132,
}

# Resistance and reactance [ohm/km per circuit] for common conductors.
# Sources: standard ACSR/AAAC/ACCC manufacturer data.
CONDUCTOR_PARAMS: dict[str, dict] = {
    "twin finch":         {"r": 0.0260, "x": 0.3210, "b": 3.40e-6},   # 2x Finch parallel
    "quad finch":         {"r": 0.0130, "x": 0.2820, "b": 6.80e-6},   # 4x Finch parallel
    "quad acsr finch":    {"r": 0.0130, "x": 0.2820, "b": 6.80e-6},
    "quad accc finch":    {"r": 0.0110, "x": 0.2750, "b": 6.80e-6},
    "ll-quad acsr finch": {"r": 0.0130, "x": 0.2820, "b": 6.80e-6},
    "single acsr finch":  {"r": 0.0519, "x": 0.3986, "b": 2.86e-6},
    "quad egret":         {"r": 0.0162, "x": 0.2970, "b": 6.10e-6},
    "twin mallard":       {"r": 0.0339, "x": 0.3270, "b": 3.52e-6},   # 2x Mallard
    "quad mallard":       {"r": 0.0169, "x": 0.2900, "b": 7.04e-6},
    "twin accc mallard":  {"r": 0.0280, "x": 0.3180, "b": 3.52e-6},
    "twin aaac":          {"r": 0.0320, "x": 0.3300, "b": 3.30e-6},
    "mallard":            {"r": 0.0677, "x": 0.4016, "b": 2.84e-6},   # single Mallard
    "accc mallard":       {"r": 0.0550, "x": 0.3900, "b": 2.84e-6},
    "grosbeak":           {"r": 0.0852, "x": 0.4097, "b": 2.73e-6},
    "accc grosbeak":      {"r": 0.0650, "x": 0.3900, "b": 2.73e-6},
    "acsr grosbeak":      {"r": 0.0852, "x": 0.4097, "b": 2.73e-6},
    "aaac":               {"r": 0.0640, "x": 0.4000, "b": 2.73e-6},
    "accc aaac":          {"r": 0.0520, "x": 0.3800, "b": 2.73e-6},
    "hawk":               {"r": 0.1200, "x": 0.4200, "b": 2.60e-6},
    "accc hawk":          {"r": 0.0900, "x": 0.4000, "b": 2.60e-6},
    "xlpe":               {"r": 0.0150, "x": 0.1200, "b": 2.00e-5},   # underground cable
    "cu.cable":           {"r": 0.0750, "x": 0.1300, "b": 2.00e-5},
}

DEFAULT_PARAMS = {"r": 0.0852, "x": 0.4097, "b": 2.73e-6}  # Grosbeak fallback


def _conductor_params(conductor_name: str) -> dict:
    key = conductor_name.lower().strip()
    # Try progressively shorter prefixes
    if key in CONDUCTOR_PARAMS:
        return CONDUCTOR_PARAMS[key]
    for k, v in CONDUCTOR_PARAMS.items():
        if k in key or key in k:
            return v
    return DEFAULT_PARAMS


def _parse_circuit_count(val: str) -> int:
    v = val.strip().lower()
    if v == "single":
        return 1
    if v == "double":
        return 2
    if v == "four":
        return 4
    try:
        return int(float(v))
    except ValueError:
        return 1


# ---------------------------------------------------------------------------
# 3a. Parse line name → (bus0_raw, bus1_raw, is_lilo, lilo_endpoint)
# ---------------------------------------------------------------------------

LILO_RE = re.compile(
    r"LILO\s+of\s+(.+?)\s+(?:line\s+)?(?:at|to)\s+(.+)",
    re.IGNORECASE,
)
LILO_RE2 = re.compile(
    r"LILI?\s+of\s+(.+?)\s+(?:line\s+)?(?:at|to)\s+(.+)",
    re.IGNORECASE,
)


def parse_line_name(line_name: str):
    """
    Return (bus0_raw, bus1_raw, is_lilo) where:
      - for regular lines  : bus0/bus1 are the two endpoints
      - for LILO lines     : bus0 is the inserted substation ("at X"),
                             bus1 is the first endpoint of the tapped line
    """
    name = line_name.strip()

    # Special one-off patterns that don't parse well generically
    # "Bera PP to Baghabari -Ishwardi line" → Baghabari
    m_bera = re.match(r"Bera PP to (.+?)\s*[-–]", name, re.IGNORECASE)
    if m_bera:
        return m_bera.group(1).strip(), "Baghabari", False

    # "T-connection from X to Y" or "T-connection from X toY" (no space before Y)
    m_t = re.search(r"[Tt]-?connection\s+from\s+(.+?)\s+to\s*(\S.*)", name, re.IGNORECASE)
    if m_t:
        return m_t.group(1).strip(), m_t.group(2).strip(), False

    # "Aminbazar 400/132kV transformer …" → internal transformer
    if re.match(r"Aminbazar\s+400/132", name, re.IGNORECASE):
        return "Aminbazar", "Aminbazar", False

    # "Ghorasal GIS2AIS …" / similar internal busbar connections
    m_gis = re.match(r"(\w+)\s+GIS2AIS", name, re.IGNORECASE)
    if m_gis:
        sub = m_gis.group(1).strip()
        return sub, sub, False

    # "Sonargaon S/S to Megnaghat Rental PP" → Sonargaon → Meghnaghat
    m_ss_to = re.match(r"(.+?)\s+S/S\s+to\s+(.+?)(?:\s+(?:PP|Rental\s+PP))?$", name, re.IGNORECASE)
    if m_ss_to:
        return m_ss_to.group(1).strip(), m_ss_to.group(2).strip(), False

    # LILO / LILI patterns
    for pattern in (LILO_RE, LILO_RE2):
        m = pattern.search(name)
        if m:
            tapped_line = m.group(1).strip()
            insert_sub  = m.group(2).strip()
            # Clean up stray text after the insertion-point name
            insert_sub = re.sub(r"\s+(substation|s/s|grid|ss)$", "", insert_sub, flags=re.IGNORECASE).strip()
            # Take the first endpoint of the tapped line
            parts = re.split(r"\s*[-–]\s*", tapped_line, maxsplit=1)
            first_end = parts[0].strip()
            # Remove noise tokens from first_end
            first_end = re.sub(r"^(the\s+)", "", first_end, flags=re.IGNORECASE).strip()
            return insert_sub, first_end, True

    # --- Handle "X PP to Y SS" / "X to Y" generator-substation connections ---
    # e.g. "Noapara PP to Noapara Ss" → both endpoints are same substation
    m_pp = re.match(
        r"(.+?)\s+PP\s+to\s+(.+?)\s+[Ss][Ss]$", name, re.IGNORECASE
    )
    if m_pp:
        return m_pp.group(1).strip(), m_pp.group(2).strip(), False

    m_pp2 = re.match(
        r"(.+?)\s+(?:PP|S/S|to\s+Megnaghat\s+Rental\s+PP|desh\s+energy\s+PP)\s+to\s+(.+?)(?:\s+[Ss][Ss])?$",
        name, re.IGNORECASE,
    )
    if m_pp2:
        return m_pp2.group(1).strip(), m_pp2.group(2).strip(), False

    # "Shiddhirganj to Siddhirganj Dutch Bangla PP" → both are Siddhirganj
    m_to = re.match(r"(\w+)\s+to\s+(\w+)\s+", name)
    if m_to:
        return m_to.group(1).strip(), m_to.group(2).strip(), False

    # Remove whole-name qualifiers (not part of bus identity)
    # NOTE: Do NOT strip "(N)", "(S)", "(C)", "(W)" — those ARE bus-name suffixes.
    qualifier_pats = [
        r"\s+HVDC$",
        r"^HVDC\s+",
        r"\s+\(With\s+River\)$",
        r"\s+\(O/H\)$",
        r"\s+\(U/H\)$",
        r"\s+\(U/G\)$",
        r"\s+\(overhead\)$",
        r"\s+\(UG\)$",
        r"\s+River\s+Crossing$",
        r"\s+\d+kV$",
        r"\s+GIS2AIS.*$",
        r"\s+(double circuit|3rd circuit|single circuit).*$",
        r"\s*\(Partial\)$",
        r"\s*\(Section onward\)$",
        r"\s*\(Upto river crossing\)$",
    ]
    for pat in qualifier_pats:
        name = re.sub(pat, "", name, flags=re.IGNORECASE).strip()

    # Normalise "210 MW P/S -Haripur" → preserve the dash as separator
    name = re.sub(r"\s+\d+\s*MW\s+P/S\s*", " ", name, flags=re.IGNORECASE).strip()

    # Split on first dash/hyphen that separates two bus names.
    parts = re.split(r"\s*[-–]\s*", name, maxsplit=1)
    if len(parts) == 2:
        b0, b1 = parts[0].strip(), parts[1].strip()
        # Only strip qualifiers that are clearly not bus-name identifiers
        for q in [r"\s+LILO\s+point$", r"\s+2nd\s+circuit$", r"\s+d/c\s+line$"]:
            b0 = re.sub(q, "", b0, flags=re.IGNORECASE).strip()
            b1 = re.sub(q, "", b1, flags=re.IGNORECASE).strip()
        return b0, b1, False

    return name.strip(), "", False


# ---------------------------------------------------------------------------
# 3b. Resolve raw bus name to a canonical bus name
# ---------------------------------------------------------------------------

def build_bus_lookup(buses_df: pd.DataFrame) -> dict[str, dict[int, str]]:
    """
    Return {normalised_substation_name: {voltage: bus_name}} for fast lookups.
    """
    lookup: dict[str, dict[int, str]] = {}
    for _, row in buses_df.iterrows():
        sub  = row["substation"]
        key  = normalise_bus_name(sub)
        v    = int(row["v_nom"])
        bname = row["name"]
        lookup.setdefault(key, {})[v] = bname
        # Also index by lower-case stripped name for fuzzy matching
        lookup.setdefault(sub.lower().strip(), {})[v] = bname
    return lookup


def resolve_bus(raw: str, voltage: int, lookup: dict) -> str | None:
    """
    Given a raw endpoint name (from line CSV) and the line's voltage level,
    return the canonical bus name from the bus lookup, or None if not found.

    Resolution order:
    1. Exact name + exact voltage
    2. Exact name + any voltage (nearest level preferred)
    3. Fuzzy partial name match + exact voltage
    4. Fuzzy partial name match + any voltage
    """
    if not raw:
        return None

    canonical = normalise_bus_name(raw)
    candidates = list({canonical, raw.strip()})   # deduplicated

    # Normalise raw name a second time via aliases (catches things like "Comilla (N)")
    canonical2 = normalise_bus_name(canonical)
    candidates = list({canonical, canonical2, raw.strip()})

    # --- pass 1: exact voltage ---
    for c in candidates:
        for key in (c, c.lower()):
            if key in lookup and voltage in lookup[key]:
                return lookup[key][voltage]

    # --- pass 2: same name, nearest voltage level ---
    PREF = {400: [400, 230, 132], 230: [230, 400, 132], 132: [132, 230, 400]}
    for c in candidates:
        for key in (c, c.lower()):
            if key in lookup:
                for v in PREF.get(voltage, [voltage]):
                    if v in lookup[key]:
                        return lookup[key][v]

    # --- pass 3: fuzzy partial match, exact voltage ---
    key_lower = canonical.lower()
    for k, v_map in lookup.items():
        if k and len(k) > 4 and len(key_lower) > 4:
            if k.lower() == key_lower or key_lower in k.lower() or k.lower() in key_lower:
                if voltage in v_map:
                    return v_map[voltage]

    # --- pass 4: fuzzy + any voltage ---
    for k, v_map in lookup.items():
        if k and len(k) > 4 and len(key_lower) > 4:
            if k.lower() == key_lower or key_lower in k.lower() or k.lower() in key_lower:
                for v in PREF.get(voltage, [voltage]):
                    if v in v_map:
                        return v_map[v]

    return None


# ---------------------------------------------------------------------------
# 4.  MAIN
# ---------------------------------------------------------------------------

def main():
    # --- Buses ---
    print("Parsing substations from Grid.xlsx …")
    buses_df = parse_substations()
    print(f"  {len(buses_df)} buses found")

    out_buses = OUT_DIR / "pypsa_buses.csv"
    buses_df.to_csv(out_buses, index=False)
    print(f"  Written → {out_buses}")

    # --- Lines ---
    print("\nParsing transmission lines from grids-formatted.csv …")
    lookup = build_bus_lookup(buses_df)

    line_rows = []
    unresolved = []

    with open(DATA_DIR / "grids-formatted.csv", newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for record in reader:
            line_name  = record.get("Name of Lines", "").strip()
            route_km   = record.get("Length in Route km", "").strip()
            ckt_km     = record.get("Length in Ckt. Km", "").strip()
            n_ckt_raw  = record.get("No. of Ckt.", "").strip()
            conductor  = record.get("Conductor", "").strip()
            voltage_s  = record.get("Voltage Level", "").strip()

            if not line_name or not voltage_s:
                continue

            voltage = VOLTAGE_MAP.get(voltage_s, VOLTAGE_MAP.get(voltage_s.replace(" ", ""), None))
            if voltage is None:
                continue

            n_ckt = _parse_circuit_count(n_ckt_raw) if n_ckt_raw else 1

            # Length per circuit
            try:
                length_per_ckt = float(ckt_km) / n_ckt if ckt_km else (float(route_km) if route_km else 0.0)
            except ValueError:
                length_per_ckt = 0.0

            params = _conductor_params(conductor)

            bus0_raw, bus1_raw, is_lilo = parse_line_name(line_name)

            bus0 = resolve_bus(bus0_raw, voltage, lookup) if bus0_raw else None
            bus1 = resolve_bus(bus1_raw, voltage, lookup) if bus1_raw else None

            if not bus0 or not bus1:
                unresolved.append({
                    "line_name": line_name,
                    "voltage":   voltage,
                    "bus0_raw":  bus0_raw,
                    "bus1_raw":  bus1_raw,
                    "bus0_resolved": bus0,
                    "bus1_resolved": bus1,
                    "is_lilo":   is_lilo,
                })
                # Still emit partial row so we can see what's missing
                bus0 = bus0 or f"UNRESOLVED:{bus0_raw}"
                bus1 = bus1 or f"UNRESOLVED:{bus1_raw}"

            for ckt in range(1, n_ckt + 1):
                # Naming: <bus0>_<bus1>_TL_<ckt>  (strip voltage suffix for brevity)
                b0_short = bus0.replace(f"_{voltage}kV", "") if bus0 else bus0_raw
                b1_short = bus1.replace(f"_{voltage}kV", "") if bus1 else bus1_raw
                name = f"{b0_short}_{b1_short}_TL_{ckt}"
                if is_lilo:
                    name = "LILO_" + name

                line_rows.append({
                    "name":        name,
                    "bus0":        bus0,
                    "bus1":        bus1,
                    "length":      round(length_per_ckt, 3),
                    "r_ohm_per_km": params["r"],
                    "x_ohm_per_km": params["x"],
                    "b_siemens_per_km": params["b"],
                    "v_nom_kv":    voltage,
                    "n_circuits":  n_ckt,
                    "conductor":   conductor,
                    "is_lilo":     is_lilo,
                    "source_name": line_name,
                })

    lines_df = pd.DataFrame(line_rows)

    out_lines = OUT_DIR / "pypsa_lines.csv"
    lines_df.to_csv(out_lines, index=False)
    print(f"  {len(lines_df)} line records written → {out_lines}")

    if unresolved:
        out_unres = OUT_DIR / "unresolved_lines.csv"
        pd.DataFrame(unresolved).to_csv(out_unres, index=False)
        print(f"\n  ⚠  {len(unresolved)} lines had unresolved endpoints → {out_unres}")

    # --- Summary ---
    print("\n=== Bus summary ===")
    print(buses_df.groupby("v_nom").size().rename("count"))
    print("\n=== Line summary ===")
    print(lines_df.groupby(["v_nom_kv", "is_lilo"]).size().rename("count"))


if __name__ == "__main__":
    main()
